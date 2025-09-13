##############################################################################
# Name: test_driver.py
#
# - Inference script for test result collection
###############################################################################

from .test_parameter import *

import csv
from datetime import datetime
import pandas as pd
import ray
import numpy as np
import os
import torch

from openpyxl.workbook.properties import CalcProperties
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from taxabind_avs.satbind.clip_seg_tta import ClipSegTTA
from .model import PolicyNet
from .test_worker import TestWorker
from .test_info_surfing import ISEnv

np.seterr(invalid='raise', divide='raise')

def run_test():

    print("\n====================")
    print("Key Test Params:")
    for key, value in OPT_VARS.items():
        print(f"{GREEN}{key}{NC}: {value}")
    print("====================\n")
        
    # Create .csv file for data collection
    current_datetime = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    csv_file_name = "{}_{}.csv".format(CSV_EXPT_NAME, current_datetime)
    csv_file_path = os.path.join(LOG_PATH, csv_file_name)
    global NUM_TEST

    # Create CSV file
    if not os.path.exists(LOG_PATH):
        os.makedirs(LOG_PATH)
    if not os.path.exists(csv_file_path):
        with open(csv_file_path, mode='w') as csv_file:
            fieldnames = ['eps', 'tax', 'num_robots', 'max_dist', 'steps', 'steps_to_first_tgt', 'steps_to_mid_tgt', 'steps_to_last_tgt', 'explored', 'targets_found', 'targets_found_abs', 'targets_total', 'kmeans_k', 'tgts_gt_score', 'clip_inference_time', 'tta_time', 'success']
            writer = csv.DictWriter(csv_file, fieldnames=fieldnames)
            writer.writeheader()

    device = torch.device('cuda') if USE_GPU else torch.device('cpu')
    global_network = PolicyNet(INPUT_DIM, EMBEDDING_DIM).to(device)

    if device == 'cuda':
        checkpoint = torch.load(f'{MODEL_PATH}/{MODEL_NAME}')
    else:
        checkpoint = torch.load(f'{MODEL_PATH}/{MODEL_NAME}', map_location = torch.device('cpu'))

    global_network.load_state_dict(checkpoint['policy_model'])

    meta_agents = [Runner.remote(i) for i in range(NUM_META_AGENT)]
    weights = global_network.state_dict()
    curr_test = 0

    dist_history = []

    # NOTE: Override NUM_TEST
    if LOAD_AVS_BENCH:
        inat_search_ds = ray.get(meta_agents[0].get_clip_seg_tta.remote()).dataset
        if not inat_search_ds.sat_to_img_ids_json_is_train_dict:
            NUM_TEST = len(inat_search_ds.filtered_val_ds_by_tax)
            print("[AVS-Bench] Overriding NUM_TEST to:", NUM_TEST)

    job_list = []
    for i, meta_agent in enumerate(meta_agents):
        job_list.append(meta_agent.job.remote(weights, curr_test))
        curr_test += 1


    try:
        while len(dist_history) < curr_test:
            done_id, job_list = ray.wait(job_list)
            done_jobs = ray.get(done_id)

            for job in done_jobs:
                metrics, info = job
                dist_history.append(metrics['travel_dist'])
                if metrics['targets_found'] is not None and metrics['targets_total'] is not None:
                    try:
                        targets_found_abs = metrics['targets_found'] * metrics['targets_total']
                    except FloatingPointError as e:
                        print(f"{RED}Caught floating point error for targets_found_abs:{NC} {e}")
                        print("metrics['targets_total']: ", metrics['targets_total'] )
                        targets_found_abs = None
                else:
                    targets_found_abs = None

                # Populate CSV file
                with open(csv_file_path, mode='a') as csv_file:
                    fieldnames = ['eps', 'tax', 'num_robots', 'max_dist', 'steps', 'steps_to_first_tgt', 'steps_to_mid_tgt', 'steps_to_last_tgt', 'explored', 'targets_found', 'targets_found_abs', 'targets_total', 'kmeans_k', 'tgts_gt_score', 'clip_inference_time', 'tta_time', 'success']
                    writer = csv.DictWriter(csv_file, fieldnames=fieldnames)
                    writer.writerow({'eps': info['episode_number'], \
                                    'num_robots': NUM_ROBOTS, \
                                    'tax': metrics['tax'], \
                                    'max_dist': metrics['travel_dist'], \
                                    'steps': metrics['travel_steps'], \
                                    'steps_to_first_tgt': metrics['steps_to_first_tgt'], \
                                    'steps_to_mid_tgt': metrics['steps_to_mid_tgt'], \
                                    'steps_to_last_tgt': metrics['steps_to_last_tgt'], \
                                    'explored': metrics['explored_rate'], \
                                    'targets_found': metrics['targets_found'], \
                                    'targets_found_abs': targets_found_abs, \
                                    'targets_total': metrics['targets_total'], \
                                    'kmeans_k': metrics['kmeans_k'], \
                                    'tgts_gt_score': metrics['tgts_gt_score'], \
                                    'clip_inference_time': metrics['clip_inference_time'], \
                                    'tta_time': metrics['tta_time'], \
                                    'success': metrics['success_rate'] })

            if curr_test < NUM_TEST:
                job_list.append(meta_agents[info['id']].job.remote(weights, curr_test))
                curr_test += 1

        # Sort CSV file by episode number
        df = pd.read_csv(csv_file_path)
        sorted_df = df.sort_values(by='eps')
        sorted_df.to_csv(csv_file_path, index=False)
        populate_raw_data(csv_file_path)

        dist_history = [d for d in dist_history if d is not None]
        print('|#Total test:', NUM_TEST)
        print('|#Average length:', np.array(dist_history).mean())
        print('|#Length std:', np.array(dist_history).std())
        print("|#CSV file saved at:", csv_file_path)

    except KeyboardInterrupt:
        print("CTRL_C pressed. Killing remote workers")
        for a in meta_agents:
            ray.kill(a)
        populate_raw_data(csv_file_path)
        print("|#CSV file saved at:", csv_file_path)
        

def populate_raw_data(csv_path, template_xlsx=LOG_TEMPLATE_XLSX, output_xlsx = None):
    """
    Load a template .xlsx with formulas in other sheets, and paste CSV content
    into a sheet named 'Raw Data' (assumed empty in the template).

    Args:
        template_xlsx: Path to your template Excel file.
        csv_path: Path to CSV file with raw data.
        output_xlsx: Where to save populated file (defaults to csv basename + '_populated.xlsx').

    Returns:
        Path to saved .xlsx file.
    """
    if output_xlsx is None:
        stem, _ = os.path.splitext(csv_path)
        output_xlsx = stem + "_summarized.xlsx"

    df = pd.read_csv(csv_path)
    wb = load_workbook(template_xlsx)
    if "Raw Data" not in wb.sheetnames:
        raise ValueError("Template must contain a sheet named 'Raw Data'.")

    ws = wb["Raw Data"]
    if ws.max_row > 0:
        ws.delete_rows(1, ws.max_row)

    # Write DataFrame (including headers)
    for r, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
        for c, val in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=val)

    # Force Excel to recalc all formulas in Summary/etc.
    wb.calculation = CalcProperties(calcId=0, fullCalcOnLoad=True)

    wb.save(output_xlsx)
    print("|#Populated Excel file saved at:", output_xlsx)
    return output_xlsx


@ray.remote(num_cpus=1, num_gpus=NUM_GPU/NUM_META_AGENT)
class Runner(object):
    def __init__(self, meta_agent_id):
        self.meta_agent_id = meta_agent_id
        self.device = torch.device('cuda') if USE_GPU else torch.device('cpu')
        self.local_network = PolicyNet(INPUT_DIM, EMBEDDING_DIM)
        self.local_network.to(self.device)

        if LOAD_AVS_BENCH:
            self.clip_seg_tta = ClipSegTTA(
                img_dir=AVS_IMG_DIR,
                imo_dir=AVS_IMO_DIR,
                json_path=AVS_INAT_JSON_PATH,
                sat_to_img_ids_path=AVS_SAT_TO_IMG_IDS_PATH,
                sat_checkpoint_path=AVS_SAT_CHECKPOINT_PATH,
                load_pretrained_hf_ckpt=AVS_LOAD_PRETRAINED_HF_CHECKPOINT,
                blur_kernel = AVS_GAUSSIAN_BLUR_KERNEL,
                device=self.device,
                sat_to_img_ids_json_is_train_dict=False, 
                tax_to_filter_val=QUERY_TAX,
                load_model=USE_CLIP_PREDS,
                query_modality=QUERY_MODALITY,
                sound_dir = AVS_SOUND_DIR,
                sound_checkpoint_path=AVS_SOUND_CHECKPOINT_PATH,
            )
        else:
            self.clip_seg_tta = None

    def set_weights(self, weights):
        self.local_network.load_state_dict(weights)

    def do_job(self, episode_number):

        n_agent = NUM_ROBOTS
        if POLICY == "RL":
            worker = TestWorker(self.meta_agent_id, n_agent, self.local_network, episode_number, device=self.device, save_image=SAVE_GIFS, greedy=True, clip_seg_tta=self.clip_seg_tta)
            worker.work(episode_number)
        elif POLICY == "IS":
            worker = ISEnv(global_step=episode_number, shape=(NUM_COORDS_HEIGHT,NUM_COORDS_WIDTH), numAgents=n_agent, save_image=SAVE_GIFS, clip_seg_tta=self.clip_seg_tta)
            worker.is_scenario(NUM_EPS_STEPS, episode_number)

        perf_metrics = worker.perf_metrics
        return perf_metrics

    def job(self, weights, episode_number):
        print(GREEN, "\nstarting episode {} on metaAgent {}".format(episode_number, self.meta_agent_id), NC)
        # set the local weights to the global weight values from the master network
        self.set_weights(weights)

        metrics = self.do_job(episode_number)

        info = {
            "id": self.meta_agent_id,
            "episode_number": episode_number,
        }

        return metrics, info
    
    def get_clip_seg_tta(self):
        return self.clip_seg_tta


if __name__ == '__main__':
    ray.init()
    print("Welcome to VLM-Search Inference Sim!")
    for i in range(NUM_RUN):
        run_test()
